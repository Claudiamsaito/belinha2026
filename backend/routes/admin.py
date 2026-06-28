import io
from flask import Blueprint, request, jsonify, send_file
from routes.auth import requer_auth
import database as db_module

admin_bp = Blueprint("admin", __name__)


@admin_bp.route("/dashboard", methods=["GET"])
@requer_auth
def dashboard():
    periodo = request.args.get("periodo")
    dias = _periodo_para_dias(periodo)

    avaliacoes = db_module.obter_todas_avaliacoes(dias)
    stats_por_unidade = db_module.obter_todas_stats(dias)

    total = len(avaliacoes)
    media_geral = 0.0
    if stats_por_unidade:
        medias = [s["media_geral"] for s in stats_por_unidade.values() if s["media_geral"] > 0]
        media_geral = round(sum(medias) / len(medias), 1) if medias else 0

    return jsonify({
        "resumo": {
            "total_avaliacoes": total,
            "unidades_avaliadas": len(stats_por_unidade),
            "media_geral": media_geral,
        },
        "stats_por_unidade": stats_por_unidade,
        "periodo": periodo or "todos",
    })


@admin_bp.route("/export/excel", methods=["GET"])
@requer_auth
def exportar_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"erro": "openpyxl não instalado"}), 500

    periodo = request.args.get("periodo")
    dias = _periodo_para_dias(periodo)
    avaliacoes = db_module.obter_todas_avaliacoes(dias)

    wb = Workbook()
    ws = wb.active
    ws.title = "Avaliações"

    cabecalho_style = Font(bold=True, color="FFFFFF")
    cabecalho_fill = PatternFill(fill_type="solid", fgColor="2E7D32")
    cabecalho_align = Alignment(horizontal="center")

    cabecalhos = [
        "ID", "Unidade ID", "Atend. Recepção", "Tempo Espera Recepção",
        "Tempo Espera Consulta", "Infraestrutura", "Retorno Elogio",
        "Comentário", "IP", "Data",
    ]
    for col, texto in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=texto)
        cell.font = cabecalho_style
        cell.fill = cabecalho_fill
        cell.alignment = cabecalho_align

    for row_idx, av in enumerate(avaliacoes, 2):
        ws.cell(row=row_idx, column=1, value=av.get("id"))
        ws.cell(row=row_idx, column=2, value=av.get("unidade_id"))
        ws.cell(row=row_idx, column=3, value=av.get("atendimento_recepcao"))
        ws.cell(row=row_idx, column=4, value=av.get("tempo_espera_recepcao"))
        ws.cell(row=row_idx, column=5, value=av.get("tempo_espera_consulta"))
        ws.cell(row=row_idx, column=6, value=av.get("infraestrutura"))
        ws.cell(row=row_idx, column=7, value=av.get("elogio_retorno"))
        ws.cell(row=row_idx, column=8, value=av.get("comentario") or "")
        ws.cell(row=row_idx, column=9, value=av.get("ip_address") or "")
        ws.cell(row=row_idx, column=10, value=str(av.get("created_at", "")))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = f"avaliacoes_missao_saude_{periodo or 'todas'}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome_arquivo,
    )


@admin_bp.route("/audit", methods=["GET"])
@requer_auth
def audit_logs():
    limite = int(request.args.get("limite", 100))
    logs = db_module.obter_audit_logs(limite)
    return jsonify(logs)


@admin_bp.route("/audit", methods=["POST"])
@requer_auth
def registrar_audit():
    data = request.get_json()
    db_module.registrar_auditoria(
        request.admin.get("email", "?"),
        data.get("action", ""),
        data.get("details", ""),
        request.remote_addr,
    )
    return jsonify({"ok": True})


def _periodo_para_dias(periodo: str | None) -> int | None:
    mapa = {"7dias": 7, "30dias": 30, "3meses": 90}
    return mapa.get(periodo) if periodo else None
